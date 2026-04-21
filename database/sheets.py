import os
import openpyxl
from openpyxl import load_workbook

# Saves leads.xlsx in the same folder as the project
EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "leads.xlsx")

HEADERS = ["Company", "Website", "Offer", "Email", "Status", "Tracking ID", "Date", "Opened", "Replied"]


def connect_sheet():
    """
    Opens leads.xlsx if it exists, creates it with headers if it doesn't.
    Returns the workbook + sheet as a tuple — replaces the gspread sheet object.
    """
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads"
        ws.append(HEADERS)
        wb.save(EXCEL_PATH)
        print(f"Created new leads file: {EXCEL_PATH}")
    return (wb, ws)


def get_all_data(sheet):
    """
    Returns all rows as a list of dicts keyed by column header.
    Mirrors gspread's get_all_records() output so main.py needs no changes.
    """
    wb, ws = sheet
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = rows[0]
    return [
        dict(zip(headers, row))
        for row in rows[1:]
        if any(cell is not None for cell in row)  # skip empty rows
    ]


def find_existing(sheet, email):
    """
    Returns the 1-based row index of an existing email, or None.
    Row 1 is the header, data starts at row 2.
    """
    wb, ws = sheet
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[3] == email:  # column D = Email
            return idx
    return None


def save_or_update(sheet, data):
    """
    Updates the row if the email already exists, appends a new row if not.
    Saves the file to disk after every write.
    """
    wb, ws = sheet

    row_index = find_existing(sheet, data[3])

    if row_index:
        for col_idx, value in enumerate(data, start=1):
            ws.cell(row=row_index, column=col_idx, value=value)
        wb.save(EXCEL_PATH)
        return "updated"
    else:
        ws.append(data)
        wb.save(EXCEL_PATH)
        return "new"